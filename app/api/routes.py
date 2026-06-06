"""
API路由定义（V2.0 升级）
保留V1全部功能 + 新增客户发现、搜索任务管理、断点续跑等API
"""
import json
import datetime
import os
import asyncio
from typing import Optional, Dict, Set
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from sqlalchemy.orm import Session
from app.database import get_db, Customer, SearchTask
from app.services.excel_importer import parse_excel, import_customers
from app.services.website_scraper import scrape_website
from app.services.email_extractor import extract_emails_from_text
from app.services.keyword_analyzer import analyze_keywords
from app.services.deepseek_analyzer import analyze_company, generate_summary
from app.services.scoring_engine import calculate_scores
from app.services.search_task_service import run_search_task, request_stop, get_stop_flag, get_paused_tasks, resume_paused_task
from app.services.keyword_expander import expand_keywords

router = APIRouter(prefix="/api", tags=["customers"])

# ──── 全局分析状态控制 ────
_analyzing_set: Set[int] = set()
_running_tasks: Set[int] = set()


def _get_emails_list(customer: Customer) -> list:
    if not customer.emails:
        return []
    try:
        return json.loads(customer.emails)
    except (json.JSONDecodeError, TypeError):
        return [e.strip() for e in customer.emails.split(",") if e.strip()]


# ═══════════════════════════════════════════
# V1 保留功能：客户管理
# ═══════════════════════════════════════════

@router.get("/customers")
def list_customers(
    search: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    sort_by_score: Optional[str] = Query("desc"),
    db: Session = Depends(get_db),
):
    query = db.query(Customer)
    if search:
        query = query.filter(Customer.company_name.ilike(f"%{search}%"))
    if country:
        query = query.filter(Customer.country == country)
    if priority:
        query = query.filter(Customer.priority == priority.upper())
    if sort_by_score == "asc":
        query = query.order_by(Customer.total_score.asc().nullslast())
    else:
        query = query.order_by(Customer.total_score.desc().nullslast())

    customers = query.all()

    all_countries = db.query(Customer.country).distinct().filter(
        Customer.country.isnot(None), Customer.country != ""
    ).order_by(Customer.country).all()
    country_list = [c[0] for c in all_countries if c[0]]

    result = []
    for c in customers:
        emails = _get_emails_list(c)
        result.append({
            "id": c.id,
            "company_name": c.company_name,
            "website": c.website or "",
            "country": c.country or "",
            "email_count": len(emails),
            "total_score": c.total_score,
            "priority": c.priority or "-",
            "ai_summary": c.ai_summary or "",
            "discovery_source": c.discovery_source or "",
            "discovery_keyword": c.discovery_keyword or "",
            "is_analyzing": c.id in _analyzing_set,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "analyzed_at": c.analyzed_at.isoformat() if c.analyzed_at else None,
        })

    return {"customers": result, "total": len(result), "countries": country_list}


@router.get("/customers/{customer_id}")
def get_customer_detail(customer_id: int, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")

    emails = _get_emails_list(customer)

    positive_kw = {}
    if customer.positive_keywords:
        try:
            positive_kw = json.loads(customer.positive_keywords)
        except (json.JSONDecodeError, TypeError):
            pass

    negative_kw = {}
    if customer.negative_keywords:
        try:
            negative_kw = json.loads(customer.negative_keywords)
        except (json.JSONDecodeError, TypeError):
            pass

    ai_raw = {}
    if customer.ai_raw_json:
        try:
            ai_raw = json.loads(customer.ai_raw_json)
        except (json.JSONDecodeError, TypeError):
            pass

    return {
        "id": customer.id,
        "company_name": customer.company_name,
        "website": customer.website or "",
        "country": customer.country or "",
        "emails": emails,
        "website_text": customer.website_text or "",
        "positive_keywords": positive_kw,
        "negative_keywords": negative_kw,
        "industry_score": customer.industry_score,
        "project_score": customer.project_score,
        "company_type_score": customer.company_type_score,
        "country_score": customer.country_score,
        "contact_score": customer.contact_score,
        "total_score": customer.total_score,
        "priority": customer.priority or "-",
        "company_type": customer.company_type or "",
        "ai_summary": customer.ai_summary or "",
        "sales_hook": customer.sales_hook or "",
        "target_position": customer.target_position or "",
        "identified_projects": customer.identified_projects or "",
        "discovery_source": customer.discovery_source or "",
        "discovery_keyword": customer.discovery_keyword or "",
        "first_found_at": customer.first_found_at.isoformat() if customer.first_found_at else None,
        "ai_raw": ai_raw,
        "is_analyzing": customer.id in _analyzing_set,
        "created_at": customer.created_at.isoformat() if customer.created_at else None,
        "analyzed_at": customer.analyzed_at.isoformat() if customer.analyzed_at else None,
    }


@router.post("/import-excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 格式的Excel文件")
    os.makedirs("app/uploads", exist_ok=True)
    upload_path = f"app/uploads/{file.filename}"
    content = await file.read()
    with open(upload_path, "wb") as f:
        f.write(content)
    try:
        customers_data = parse_excel(upload_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析Excel失败: {str(e)}")
    if not customers_data:
        raise HTTPException(status_code=400, detail="Excel中没有找到有效数据")
    count = import_customers(customers_data)
    return {"message": f"成功导入 {count} 个客户", "total_in_file": len(customers_data), "imported": count, "skipped": len(customers_data) - count}


@router.post("/analyze/{customer_id}")
async def analyze_single(customer_id: int, db: Session = Depends(get_db)):
    if customer_id in _analyzing_set:
        raise HTTPException(status_code=400, detail="该客户正在分析中")
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    if not customer.website:
        raise HTTPException(status_code=400, detail="该客户没有官网地址")
    _analyzing_set.add(customer_id)
    try:
        website_text = await scrape_website(customer.website)
        if website_text:
            customer.website_text = website_text
            emails = extract_emails_from_text(website_text)
            email_list = list(set(emails))
            customer.emails = json.dumps(email_list, ensure_ascii=False)
            pos_hits, neg_hits = analyze_keywords(website_text)
            customer.positive_keywords = json.dumps(pos_hits, ensure_ascii=False)
            customer.negative_keywords = json.dumps(neg_hits, ensure_ascii=False)
            ai_result = await analyze_company(website_text)
            if ai_result:
                customer.ai_raw_json = json.dumps(ai_result, ensure_ascii=False)
                customer.company_type = ai_result.get("company_type", "")
                customer.sales_hook = ai_result.get("sales_hook", "")
                customer.target_position = ai_result.get("target_position", "")
                customer.identified_projects = ai_result.get("identified_projects", "")
                customer_info = {"country": customer.country or "", "company_name": customer.company_name}
                customer.ai_summary = generate_summary(ai_result, customer_info)
            scores = calculate_scores(
                website_text=website_text, positive_keywords=pos_hits,
                company_type=customer.company_type, country=customer.country, emails=email_list,
            )
            customer.industry_score = scores["industry_score"]
            customer.project_score = scores["project_score"]
            customer.company_type_score = scores["company_type_score"]
            customer.country_score = scores["country_score"]
            customer.contact_score = scores["contact_score"]
            customer.total_score = scores["total_score"]
            customer.priority = scores["priority"]
        customer.analyzed_at = datetime.datetime.utcnow()
        db.commit()
        return {"message": "分析完成", "customer_id": customer.id}
    except Exception as e:
        db.rollback()
        raise e
    finally:
        _analyzing_set.discard(customer_id)


@router.post("/analyze-all")
async def analyze_all(db: Session = Depends(get_db)):
    customers = db.query(Customer).filter(Customer.analyzed_at.is_(None)).all()
    if not customers:
        return {"message": "没有待分析的客户", "analyzed_count": 0}
    analyzed_count = 0
    for customer in customers:
        if not customer.website:
            continue
        if customer.id in _analyzing_set:
            continue
        try:
            await analyze_single(customer.id, db)
            analyzed_count += 1
        except Exception as e:
            print(f"分析 {customer.company_name} 失败: {str(e)}")
            continue
    return {"message": "批量分析完成", "analyzed_count": analyzed_count}


@router.post("/stop-analysis")
async def stop_analysis():
    request_stop()
    return {"message": "已发送停止信号，正在等待当前任务完成"}


@router.get("/analysis-status")
def get_analysis_status():
    return {
        "is_analyzing": len(_analyzing_set) > 0,
        "analyzing_ids": list(_analyzing_set),
    }


@router.get("/export-excel")
def export_excel(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    from fastapi.responses import StreamingResponse
    customers = db.query(Customer).order_by(Customer.total_score.desc().nullslast()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客户分析结果"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ["公司名称", "国家", "官网", "邮箱数量", "总分", "优先级", "公司类型", "发现来源", "发现关键词", "AI摘要", "开发切入点", "推荐联系职位", "分析时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
    for row_idx, c in enumerate(customers, 2):
        emails = _get_emails_list(c)
        row_data = [
            c.company_name, c.country or "", c.website or "", len(emails),
            c.total_score or 0, c.priority or "-", c.company_type or "",
            c.discovery_source or "", c.discovery_keyword or "",
            c.ai_summary or "", c.sales_hook or "", c.target_position or "",
            c.analyzed_at.strftime("%Y-%m-%d %H:%M") if c.analyzed_at else "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border; cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, w in enumerate([25, 15, 35, 10, 8, 8, 18, 12, 20, 35, 30, 20, 18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=customers_export.xlsx"})


@router.delete("/customers/{customer_id}")
def delete_customer(customer_id: int, db: Session = Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    db.delete(customer); db.commit()
    return {"message": "删除成功"}


@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    total = db.query(Customer).count()
    analyzed = db.query(Customer).filter(Customer.analyzed_at.isnot(None)).count()
    pending = total - analyzed
    a = db.query(Customer).filter(Customer.priority == "A").count()
    b = db.query(Customer).filter(Customer.priority == "B").count()
    c = db.query(Customer).filter(Customer.priority == "C").count()
    d = db.query(Customer).filter(Customer.priority == "D").count()
    google_count = db.query(Customer).filter(Customer.discovery_source == "Google").count()
    manual_count = db.query(Customer).filter(Customer.discovery_source.is_(None)).count()
    return {
        "total": total, "analyzed": analyzed, "pending": pending,
        "priority_distribution": {"A": a, "B": b, "C": c, "D": d},
        "discovery_stats": {"google": google_count, "manual_import": manual_count},
    }


# ═══════════════════════════════════════════
# V2.0 新增：客户发现 & 搜索任务管理
# ═══════════════════════════════════════════

@router.post("/discovery/expand-keywords")
async def api_expand_keywords(keyword: str = Query(..., description="需要扩展的基础关键词")):
    """AI扩展关键词"""
    expanded = await expand_keywords(keyword)
    if not expanded:
        expanded = [keyword]
    return {"original_keyword": keyword, "expanded_keywords": expanded}


@router.post("/discovery/search-task")
async def create_search_task(
    country: str = Query(..., description="搜索国家"),
    keyword: str = Query(..., description="搜索关键词"),
    depth: int = Query(50, description="搜索深度"),
    db: Session = Depends(get_db),
):
    """创建新的搜索发现任务"""
    task = SearchTask(
        country=country,
        keyword=keyword,
        search_depth=depth,
        status="Pending",
        created_at=datetime.datetime.utcnow(),
    )
    db.add(task)
    db.commit()
    db.refresh(task)

    # 异步启动任务
    loop = asyncio.get_event_loop()
    loop.create_task(_run_task_wrapper(task.id))

    return {"message": "搜索任务已创建", "task_id": task.id}


async def _run_task_wrapper(task_id: int):
    """异步包装器，确保任务异常不影响主进程"""
    try:
        await run_search_task(task_id)
    except Exception as e:
        print(f"搜索任务 {task_id} 异常退出: {str(e)[:200]}")


@router.get("/discovery/tasks")
def list_search_tasks(db: Session = Depends(get_db)):
    """获取所有搜索任务列表"""
    tasks = db.query(SearchTask).order_by(SearchTask.created_at.desc()).all()
    result = []
    for t in tasks:
        expanded = []
        if t.expanded_keywords:
            try:
                expanded = json.loads(t.expanded_keywords)
            except (json.JSONDecodeError, TypeError):
                pass
        result.append({
            "id": t.id,
            "country": t.country,
            "keyword": t.keyword,
            "expanded_keywords": expanded,
            "search_depth": t.search_depth,
            "status": t.status,
            "found_websites": t.found_websites or 0,
            "analyzed_companies": t.analyzed_companies or 0,
            "new_companies": t.new_companies or 0,
            "current_keyword_index": t.current_keyword_index or 0,
            "error_message": t.error_message,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "finished_at": t.finished_at.isoformat() if t.finished_at else None,
        })

    # 标记当前活跃任务（优先 Running，其次 Pending）
    active_task_id = None
    for t in tasks:
        if t.status == "Running":
            active_task_id = t.id
            break
    if active_task_id is None:
        for t in tasks:
            if t.status == "Pending":
                active_task_id = t.id
                break

    return {"tasks": result, "total": len(result), "active_task_id": active_task_id}


@router.post("/discovery/tasks/{task_id}/pause")
def pause_task(task_id: int, db: Session = Depends(get_db)):
    """暂停搜索任务"""
    request_stop()
    return {"message": "正在停止搜索任务"}


@router.post("/discovery/tasks/{task_id}/resume")
def resume_task(task_id: int, db: Session = Depends(get_db)):
    """恢复暂停的搜索任务（断点续跑）"""
    success = resume_paused_task(task_id)
    if not success:
        raise HTTPException(status_code=400, detail="任务不存在或无法恢复")
    # 重新启动
    loop = asyncio.get_event_loop()
    loop.create_task(_run_task_wrapper(task_id))
    return {"message": "任务已恢复", "task_id": task_id}


@router.delete("/discovery/tasks/{task_id}")
def delete_task(task_id: int, db: Session = Depends(get_db)):
    """删除搜索任务"""
    task = db.query(SearchTask).filter(SearchTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    db.delete(task)
    db.commit()
    return {"message": "删除成功"}


@router.get("/discovery/paused-tasks")
def list_paused_tasks(db: Session = Depends(get_db)):
    """获取所有暂停的任务（用于断点续跑）"""
    tasks = get_paused_tasks(db)
    result = []
    for t in tasks:
        result.append({
            "id": t.id,
            "country": t.country,
            "keyword": t.keyword,
            "status": t.status,
            "current_keyword_index": t.current_keyword_index or 0,
        })
    return {"paused_tasks": result}


@router.get("/discovery/discovered-customers")
def list_discovered_customers(
    search: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None, description="按发现关键词筛选"),
    sort_by_score: Optional[str] = Query("desc"),
    db: Session = Depends(get_db),
):
    """获取通过Google发现的公司列表"""
    query = db.query(Customer).filter(Customer.discovery_source == "Google")
    if search:
        query = query.filter(Customer.company_name.ilike(f"%{search}%"))
    if country:
        query = query.filter(Customer.country == country)
    if priority:
        query = query.filter(Customer.priority == priority.upper())
    if keyword:
        query = query.filter(Customer.discovery_keyword.ilike(f"%{keyword}%"))
    if sort_by_score == "asc":
        query = query.order_by(Customer.total_score.asc().nullslast())
    else:
        query = query.order_by(Customer.total_score.desc().nullslast())

    customers = query.all()
    all_countries = db.query(Customer.country).distinct().filter(
        Customer.country.isnot(None), Customer.country != "",
        Customer.discovery_source == "Google",
    ).order_by(Customer.country).all()
    country_list = [c[0] for c in all_countries if c[0]]

    all_keywords = db.query(Customer.discovery_keyword).distinct().filter(
        Customer.discovery_keyword.isnot(None), Customer.discovery_keyword != "",
        Customer.discovery_source == "Google",
    ).all()
    keyword_list = list(set(k[0] for k in all_keywords if k[0]))

    result = []
    for c in customers:
        emails = _get_emails_list(c)
        result.append({
            "id": c.id,
            "company_name": c.company_name,
            "website": c.website or "",
            "country": c.country or "",
            "email_count": len(emails),
            "total_score": c.total_score,
            "priority": c.priority or "-",
            "discovery_keyword": c.discovery_keyword or "",
            "ai_summary": c.ai_summary or "",
            "created_at": c.created_at.isoformat() if c.created_at else None,
        })

    return {"customers": result, "total": len(result), "countries": country_list, "keywords": keyword_list}
